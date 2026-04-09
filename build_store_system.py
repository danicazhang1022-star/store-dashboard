#!/usr/bin/env python3
"""
扭捏niunie 线下店铺智能管理系统
包含：产品库 / 每日销售 / 库存管理 / 入库记录 / 薪资计算 / 月度汇总
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

SOURCE_FILE = '/Users/danica/Downloads/店铺商品资料_20260407004004_152895639_1.xlsx'
OUTPUT_FILE = '/Users/danica/WorkBuddy/20260407210337/店铺智能管理系统.xlsx'

# ─── 颜色常量 ───
C_HEADER_BG  = 'FF3949AB'   # 深紫蓝 表头
C_HEADER_FG  = 'FFFFFFFF'   # 白字
C_TITLE_BG   = 'FFE8EAF6'   # 淡紫 大标题
C_INPUT_BG   = 'FFFFF9C4'   # 淡黄 输入框
C_FORMULA_BG = 'FFE8F5E9'   # 淡绿 公式结果
C_ALT_ROW    = 'FFF5F5F5'   # 浅灰 间隔行
C_ORANGE     = 'FFFF6F00'   # 橙色 强调
C_RED_BG     = 'FFFFCDD2'   # 浅红 警告
C_BLUE_BG    = 'FFE3F2FD'   # 浅蓝 说明
C_GOLD_BG    = 'FFFFF8E1'   # 金黄 提示
C_PURPLE_BG  = 'FFF3E5F5'   # 淡紫 参数区
C_GREEN_DARK = 'FF1B5E20'   # 深绿 金额
C_SECTION_BG = 'FFE0F2F1'   # 段落背景


def thin_border():
    s = Side(style='thin', color='FFBDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    s = Side(style='medium', color='FF9E9E9E')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(): return Font(name='Arial', bold=True, color=C_HEADER_FG, size=10)
def body_font(): return Font(name='Arial', size=10)
def note_font(): return Font(name='Arial', italic=True, size=9, color='FF757575')

def style_header(cell, text, bg=C_HEADER_BG, size=10):
    cell.value = text
    cell.font = Font(name='Arial', bold=True, color=C_HEADER_FG, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()

def style_input(cell, align='center'):
    cell.fill = PatternFill('solid', start_color=C_INPUT_BG)
    cell.font = Font(name='Arial', size=10, color='FF1A237E', bold=True)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin_border()

def style_formula(cell, align='center', bold=False, color='FF37474F'):
    cell.fill = PatternFill('solid', start_color=C_FORMULA_BG)
    cell.font = Font(name='Arial', size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin_border()

def set_title(ws, text, merge_to, row=1, bg=C_TITLE_BG, size=13):
    ws.merge_cells(f'A{row}:{merge_to}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name='Arial', bold=True, size=size, color='FF303F9F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=bg)
    ws.row_dimensions[row].height = 28

def set_note(ws, text, merge_to, row=2, bg=C_GOLD_BG):
    ws.merge_cells(f'A{row}:{merge_to}{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = note_font()
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', start_color=bg)
    ws.row_dimensions[row].height = 18

def draw_headers(ws, headers, widths, row=3):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i)
        style_header(c, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


# ══════════════════════════════════════════════════════════════════════
#  Sheet 1 - 产品库
# ══════════════════════════════════════════════════════════════════════
def build_product_sheet(wb, df):
    ws = wb.create_sheet('📦产品库')

    set_title(ws, '📦 产品库  |  扭捏niunie 线下店铺', 'K', row=1)
    set_note(ws, '💡 点击图片列链接可查看商品图片 | 支持按货号/名称/颜色/尺码筛选', 'K', row=2)

    headers = ['序号','货号(SKU)','款式编码','商品名称','颜色','尺码','售价(¥)','库存','图片链接','淘宝链接']
    widths  = [5, 14, 12, 38, 20, 7, 12, 8, 44, 44]
    draw_headers(ws, headers, widths, row=3)

    def split_cs(val):
        if pd.isna(val): return '', ''
        p = str(val).split(';')
        return p[0].strip(), (p[1].strip() if len(p) > 1 else '')

    df = df.copy()
    df[['颜色','尺码']] = df['线上颜色规格'].apply(lambda x: pd.Series(split_cs(x)))

    for ri, (_, row) in enumerate(df.iterrows(), 4):
        ws.row_dimensions[ri].height = 17
        bg = C_ALT_ROW if ri % 2 == 0 else 'FFFFFFFF'
        fill = PatternFill('solid', start_color=bg)
        values = [
            ri-3,
            str(row['原始商品编码']) if pd.notna(row['原始商品编码']) else '',
            str(row['线上款式编码']) if pd.notna(row['线上款式编码']) else '',
            str(row['线上商品名称']) if pd.notna(row['线上商品名称']) else '',
            row['颜色'], row['尺码'],
            row['店铺售价'] if pd.notna(row['店铺售价']) else '',
            int(row['店铺库存']) if pd.notna(row['店铺库存']) else 0,
            str(row['图片']) if pd.notna(row['图片']) else '',
            str(row['线上链接']) if pd.notna(row['线上链接']) else '',
        ]
        aligns = ['center','center','center','left','left','center','center','center','left','left']
        for ci, (val, aln) in enumerate(zip(values, aligns), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal=aln, vertical='center')
            c.border = thin_border()
            c.fill = fill
            if ci == 7 and isinstance(val, (int, float)):
                c.number_format = '¥#,##0'
            if ci == 8 and isinstance(val, (int, float)) and val == 0:
                c.fill = PatternFill('solid', start_color=C_RED_BG)
        # 超链接列
        for ci in [9, 10]:
            c = ws.cell(row=ri, column=ci)
            url = c.value
            if url and str(url).startswith('http'):
                c.hyperlink = url
                c.font = Font(name='Arial', size=9, color='FF1565C0', underline='single')

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:J{len(df)+3}'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 2 - 每日销售开单
# ══════════════════════════════════════════════════════════════════════
def build_daily_sales_sheet(wb, df):
    ws = wb.create_sheet('🧾每日销售')

    set_title(ws, '🧾 每日销售开单  |  输入货号自动带出商品信息', 'M', row=1)

    # 信息区第2行
    ws.row_dimensions[2].height = 24
    info_labels = {
        'A2': '日期：', 'C2': '当班人员：', 'F2': '人数：',
        'H2': 'POS核对(¥)：', 'K2': '系统合计(¥)：', 'M2': '差异(¥)：'
    }
    for addr, val in info_labels.items():
        c = ws[addr]
        c.value = val
        c.font = Font(name='Arial', bold=True, size=10, color='FF37474F')
        c.alignment = Alignment(horizontal='right', vertical='center')

    # 输入框
    ws['B2'].value = '2026-04-07'; ws['B2'].number_format = 'YYYY-MM-DD'; style_input(ws['B2'])
    ws['D2'].value = '店员A'; style_input(ws['D2'])  # 当班人员姓名（逗号分隔多人）
    ws['G2'].value = 1; style_input(ws['G2'])  # 当班人数
    ws['J2'].value = 0; ws['J2'].number_format = '¥#,##0'; style_input(ws['J2'])  # POS
    ws['L2'].value = '=SUM(K5:K504)'; ws['L2'].number_format = '¥#,##0'; style_formula(ws['L2'], bold=True, color=C_GREEN_DARK)
    ws['N2'].value = '=J2-L2'; ws['N2'].number_format = '¥#,##0'; style_formula(ws['N2'], bold=True, color='FFBF360C')

    set_note(ws, '💡 在【货号】列输入货号 → 商品信息自动填入 → 填【实收金额】→ 每单记录成交人，方便大单奖励核算', 'N', row=3)

    # 表头第4行
    headers = ['序号','货号(SKU)','商品名称','颜色','尺码','标价(¥)','实收(¥)','数量','小计(¥)','成交人','大单奖励(¥)','备注']
    widths  = [5, 14, 30, 16, 7, 11, 11, 7, 11, 10, 12, 16]
    # 列数13，列K=系统合计已用，这里从A开始重排
    # 实际列：A序号 B货号 C商品名 D颜色 E尺码 F标价 G实收 H数量 I小计 J成交人 K大单奖励 L备注
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=4, column=i)
        style_header(c, h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30

    # 数据行 5~504
    for r in range(5, 505):
        ws.row_dimensions[r].height = 18
        bg = C_ALT_ROW if r % 2 == 0 else 'FFFFFFFF'

        # A 序号
        c = ws.cell(row=r, column=1)
        c.value = f'=IF(B{r}="","",ROW()-4)'
        c.font = Font(name='Arial', size=9, color='FF9E9E9E')
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border()
        c.fill = PatternFill('solid', start_color=bg)

        # B 货号（输入）
        c = ws.cell(row=r, column=2)
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)
        c.font = Font(name='Arial', size=10, color='FF1A237E', bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border()

        # C 商品名称 VLOOKUP
        c = ws.cell(row=r, column=3)
        c.value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},\'📦产品库\'!$B$4:$D$9999,2,0),"⚠️未找到"))'
        style_formula(c, align='left'); c.font = Font(name='Arial', size=9)

        # D 颜色 VLOOKUP
        c = ws.cell(row=r, column=4)
        c.value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},\'📦产品库\'!$B$4:$E$9999,4,0),""))'
        style_formula(c); c.font = Font(name='Arial', size=9)

        # E 尺码 VLOOKUP
        c = ws.cell(row=r, column=5)
        c.value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},\'📦产品库\'!$B$4:$F$9999,5,0),""))'
        style_formula(c); c.font = Font(name='Arial', size=9)

        # F 标价 VLOOKUP
        c = ws.cell(row=r, column=6)
        c.value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},\'📦产品库\'!$B$4:$G$9999,6,0),""))'
        c.number_format = '¥#,##0'; style_formula(c)
        c.font = Font(name='Arial', size=10, color='FF37474F')

        # G 实收金额（手动输入）
        c = ws.cell(row=r, column=7)
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)
        c.number_format = '¥#,##0'
        c.font = Font(name='Arial', size=10, color='FF1A237E', bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border()

        # H 数量（默认1，可改）
        c = ws.cell(row=r, column=8)
        c.value = f'=IF(B{r}="","",1)'
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)
        c.font = Font(name='Arial', size=10); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()

        # I 小计（公式）
        c = ws.cell(row=r, column=9)
        c.value = f'=IF(B{r}="","",IF(G{r}<>"",G{r}*H{r},F{r}*H{r}))'
        c.number_format = '¥#,##0'
        style_formula(c, bold=True, color=C_GREEN_DARK)

        # J 成交人（手动，记录姓名用于大单奖励）
        c = ws.cell(row=r, column=10)
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)
        c.font = Font(name='Arial', size=10, color='FF6A1B9A')
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border()

        # K 大单奖励（公式：单笔≥2000+20，≥1000+10，每日最多2单在汇总里控制）
        c = ws.cell(row=r, column=11)
        c.value = f'=IF(B{r}="","",IF(G{r}>=2000,20,IF(G{r}>=1000,10,0)))'
        c.number_format = '¥#,##0'
        style_formula(c, color='FFBF360C', bold=True)
        c.font = Font(name='Arial', size=10, bold=True, color='FFBF360C')

        # L 备注
        c = ws.cell(row=r, column=12)
        c.fill = PatternFill('solid', start_color=bg); c.border = thin_border()

    # 汇总行 505
    r = 505
    ws.row_dimensions[r].height = 26
    ws.merge_cells(f'A{r}:F{r}')
    c = ws[f'A{r}']; c.value = '当日合计'
    c.font = Font(name='Arial', bold=True, size=11, color=C_HEADER_FG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_HEADER_BG)

    total_qty = ws[f'H{r}']
    total_qty.value = f'=SUM(H5:H504)'; total_qty.number_format = '0件'
    total_qty.font = Font(name='Arial', bold=True, size=11, color=C_HEADER_FG)
    total_qty.alignment = Alignment(horizontal='center', vertical='center')
    total_qty.fill = PatternFill('solid', start_color=C_HEADER_BG); total_qty.border = thin_border()

    total_amt = ws[f'I{r}']
    total_amt.value = f'=SUM(I5:I504)'; total_amt.number_format = '¥#,##0'
    total_amt.font = Font(name='Arial', bold=True, size=12, color='FFFFFF00')
    total_amt.alignment = Alignment(horizontal='center', vertical='center')
    total_amt.fill = PatternFill('solid', start_color='FF1A237E'); total_amt.border = thin_border()

    big_bonus = ws[f'K{r}']
    big_bonus.value = f'=MIN(SUM(K5:K504),40)'  # 每日大单奖励上限2单=最多40元
    big_bonus.number_format = '¥#,##0'
    big_bonus.font = Font(name='Arial', bold=True, size=11, color='FFBF360C')
    big_bonus.alignment = Alignment(horizontal='center', vertical='center')
    big_bonus.fill = PatternFill('solid', start_color='FFFFF3E0'); big_bonus.border = thin_border()

    # 列宽补充
    for col_letter, width in [('M', 14), ('N', 12)]:
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = 'A4:L4'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 3 - 薪资计算（核心）
# ══════════════════════════════════════════════════════════════════════
def build_salary_sheet(wb):
    ws = wb.create_sheet('💰薪资计算')

    set_title(ws, '💰 薪资计算  |  日薪阶梯 + 阶梯提成 + 大单奖励', 'L', row=1)

    # ─── 参数说明区 ───
    ws.row_dimensions[2].height = 20
    ws.merge_cells('A2:L2')
    c = ws['A2']; c.value = '📌 参数区（黄色格子可修改）  |  蓝色格子为自动计算结果，请勿手动填写'
    c.font = note_font(); c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', start_color=C_GOLD_BG)

    # 参数标题
    param_title_row = 3
    ws.row_dimensions[param_title_row].height = 22
    ws.merge_cells(f'A{param_title_row}:F{param_title_row}')
    c = ws[f'A{param_title_row}']; c.value = '── 薪资参数设置 ──'
    c.font = Font(name='Arial', bold=True, size=11, color='FF37474F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_SECTION_BG)

    ws.merge_cells(f'G{param_title_row}:L{param_title_row}')
    c = ws[f'G{param_title_row}']; c.value = '── 提成阶梯（按当日店铺总销售额） ──'
    c.font = Font(name='Arial', bold=True, size=11, color='FF37474F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color='FFECE1FD')

    # 左侧：日薪参数
    params_left = [
        (4, '统计月份', '2026-04', 'YYYY-MM', False),
        (5, '店员A姓名', '店员A', '@', False),
        (6, '店员B姓名', '店员B', '@', False),
        (7, '新手期日薪(¥)', 260, '¥#,##0', False),
        (8, '稳定期日薪(¥)', 280, '¥#,##0', False),
        (9, '熟练期日薪(¥)', 300, '¥#,##0', False),
        (10, '大单奖励≥1000(¥)', 10, '¥#,##0', False),
        (11, '大单奖励≥2000(¥)', 20, '¥#,##0', False),
        (12, '每日大单上限(笔)', 2, '0', False),
    ]
    for row, label, default, fmt, _ in params_left:
        ws.row_dimensions[row].height = 22
        c_lbl = ws.cell(row=row, column=1, value=label + '：')
        c_lbl.font = Font(name='Arial', bold=True, size=10)
        c_lbl.alignment = Alignment(horizontal='right', vertical='center')
        c_lbl.fill = PatternFill('solid', start_color=C_PURPLE_BG)
        c_lbl.border = thin_border()

        c_val = ws.cell(row=row, column=2, value=default)
        c_val.number_format = fmt
        style_input(c_val)

    ws.merge_cells('C4:F12')
    info = ws['C4']
    info.value = (
        '日薪说明：\n'
        '• 新手期（入职~2周）：¥260/天\n'
        '• 稳定期（满2周后）：¥280/天\n'
        '• 熟练期（稳定达标满2周）：¥300/天\n\n'
        '达标标准：\n'
        '• 周中日均 ≥ 2200–2600\n'
        '• 周末/活动日 ≥ 6500–8000\n\n'
        '经营底线：\n'
        '• 周中持续低于1200需复盘\n'
        '• 周末持续低于4000需复盘'
    )
    info.font = Font(name='Arial', size=9, color='FF37474F')
    info.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    info.fill = PatternFill('solid', start_color='FFF8F9FA')
    info.border = thin_border()

    # 右侧：提成阶梯表
    tier_headers = ['销售区间下限(¥)', '销售区间上限(¥)', '提成比例', '说明']
    tier_data = [
        (0, 3000, 0.02, '0–3000'),
        (3000, 6000, 0.025, '3001–6000'),
        (6000, 10000, 0.03, '6001–10000'),
        (10000, 9999999, 0.04, '10000以上'),
    ]
    for ci, h in enumerate(tier_headers, 7):
        c = ws.cell(row=4, column=ci); style_header(c, h, bg='FF5C35CC')
    for ri, (lo, hi, rate, desc) in enumerate(tier_data, 5):
        ws.row_dimensions[ri].height = 22
        vals = [lo, hi if hi < 9999999 else '无上限', rate, desc]
        fmts = ['¥#,##0', '¥#,##0', '0.0%', '@']
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 7):
            c = ws.cell(row=ri, column=ci, value=v)
            c.number_format = fmt
            c.font = Font(name='Arial', size=10, bold=(ci==9))
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
            c.fill = PatternFill('solid', start_color=C_FORMULA_BG if ci == 9 else 'FFFFFFFF')

    ws.merge_cells('G9:L12')
    tip = ws['G9']
    tip.value = (
        '提成分配规则：\n'
        '• 一人当班：提成全部归个人\n'
        '• 两人当班：提成池平均分配\n'
        '• 我(老板)和小何不参与提成\n\n'
        '大单奖励（归成交人，不平分）：\n'
        '• 单笔 ≥1000：+¥10\n'
        '• 单笔 ≥2000：+¥20（取最高档）\n'
        '• 每日最多计2笔大单'
    )
    tip.font = Font(name='Arial', size=9, color='FF37474F')
    tip.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    tip.fill = PatternFill('solid', start_color='FFFCE4EC')
    tip.border = thin_border()

    # 列宽
    for col, width in [('A',20),('B',14),('C',12),('D',12),('E',12),('F',12),
                        ('G',18),('H',18),('I',12),('J',18),('K',16),('L',16)]:
        ws.column_dimensions[col].width = width

    # ─── 每日明细记录表 ───
    detail_title_row = 14
    ws.row_dimensions[detail_title_row].height = 22
    ws.merge_cells(f'A{detail_title_row}:L{detail_title_row}')
    c = ws[f'A{detail_title_row}']; c.value = '── 每日销售明细（手动录入当日数据） ──'
    c.font = Font(name='Arial', bold=True, size=11, color='FF37474F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_SECTION_BG)

    detail_headers = [
        '日期', '星期', '当班\n人员', '当班\n人数',
        '店铺\n总销售额(¥)', '提成池(¥)',
        '人均\n提成(¥)',
        '店员A\n大单奖励(¥)', '店员B\n大单奖励(¥)',
        '店员A\n日薪档(¥)', '店员B\n日薪档(¥)',
        '备注/达标情况'
    ]
    detail_widths = [13,8,14,8,16,14,12,14,14,13,13,20]
    DETAIL_HDR = 15
    for ci, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
        c = ws.cell(row=DETAIL_HDR, column=ci); style_header(c, h, bg='FF283593')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[DETAIL_HDR].height = 34

    DATA_START = 16
    for r in range(DATA_START, DATA_START + 31):
        ws.row_dimensions[r].height = 20
        bg = C_ALT_ROW if r % 2 == 0 else 'FFFFFFFF'

        # A 日期
        c = ws.cell(row=r, column=1); c.number_format = 'MM/DD'
        style_input(c); c.font = Font(name='Arial', size=10, color='FF1A237E')

        # B 星期（公式）
        c = ws.cell(row=r, column=2)
        c.value = f'=IF(A{r}="","",TEXT(A{r},"AAA"))'
        style_formula(c); c.font = Font(name='Arial', size=9)

        # C 当班人员（输入，可填：店员A、店员B、店员A,店员B）
        c = ws.cell(row=r, column=3); style_input(c)
        c.font = Font(name='Arial', size=10, color='FF6A1B9A')

        # D 当班人数（公式：逗号+1计数，或手动）
        c = ws.cell(row=r, column=4)
        c.value = f'=IF(C{r}="","",IF(ISERROR(FIND(",",C{r})),1,LEN(C{r})-LEN(SUBSTITUTE(C{r},",",""))+1))'
        style_formula(c); c.font = Font(name='Arial', size=10, bold=True)

        # E 店铺总销售额（手动）
        c = ws.cell(row=r, column=5)
        c.number_format = '¥#,##0'; style_input(c)
        c.font = Font(name='Arial', size=11, color='FF1A237E', bold=True)

        # F 提成池（阶梯公式）
        # IFS：按区间算提成
        c = ws.cell(row=r, column=6)
        c.value = (
            f'=IF(E{r}="","",IF(E{r}<=0,0,'
            f'IF(E{r}<=3000,E{r}*0.02,'
            f'IF(E{r}<=6000,E{r}*0.025,'
            f'IF(E{r}<=10000,E{r}*0.03,'
            f'E{r}*0.04)))))'
        )
        c.number_format = '¥#,##0.00'
        style_formula(c, bold=True, color=C_GREEN_DARK)
        c.font = Font(name='Arial', size=10, bold=True, color=C_GREEN_DARK)

        # G 人均提成（提成池/人数）
        c = ws.cell(row=r, column=7)
        c.value = f'=IF(OR(E{r}="",D{r}=""),"",IFERROR(F{r}/D{r},F{r}))'
        c.number_format = '¥#,##0.00'
        style_formula(c, color='FF1565C0', bold=True)
        c.font = Font(name='Arial', size=10, bold=True, color='FF1565C0')

        # H 店员A大单奖励（手动填入，来自每日销售表统计）
        c = ws.cell(row=r, column=8); style_input(c)
        c.number_format = '¥#,##0'; c.font = Font(name='Arial', size=10, color='FFBF360C')

        # I 店员B大单奖励（手动）
        c = ws.cell(row=r, column=9); style_input(c)
        c.number_format = '¥#,##0'; c.font = Font(name='Arial', size=10, color='FFBF360C')

        # J 店员A日薪档（输入260/280/300）
        c = ws.cell(row=r, column=10); style_input(c)
        c.number_format = '¥#,##0'; c.value = 260
        c.font = Font(name='Arial', size=10, color='FF1A237E', bold=True)

        # K 店员B日薪档
        c = ws.cell(row=r, column=11); style_input(c)
        c.number_format = '¥#,##0'; c.value = 260
        c.font = Font(name='Arial', size=10, color='FF1A237E', bold=True)

        # L 备注
        c = ws.cell(row=r, column=12)
        c.fill = PatternFill('solid', start_color=bg); c.border = thin_border()
        c.font = Font(name='Arial', size=9)

    # ─── 月末汇总 ───
    END = DATA_START + 31
    ws.row_dimensions[END].height = 22
    ws.merge_cells(f'A{END}:L{END}')
    c = ws[f'A{END}']; c.value = '── 月末汇总 ──'
    c.font = Font(name='Arial', bold=True, size=11, color='FF37474F')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color='FFFCE4EC')

    sum_row = END + 1
    ws.row_dimensions[sum_row].height = 24

    # 月总销售额
    ws.cell(row=sum_row, column=1, value='月总销售额').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=sum_row, column=1).border = thin_border()
    ws.cell(row=sum_row, column=1).fill = PatternFill('solid', start_color=C_PURPLE_BG)
    c = ws.cell(row=sum_row, column=2)
    c.value = f'=SUM(E{DATA_START}:E{DATA_START+30})'; c.number_format = '¥#,##0'
    style_formula(c, bold=True, color='FF1A237E')

    # 月提成池
    ws.cell(row=sum_row, column=3, value='月提成池').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=sum_row, column=3).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=sum_row, column=3).border = thin_border()
    ws.cell(row=sum_row, column=3).fill = PatternFill('solid', start_color=C_PURPLE_BG)
    c = ws.cell(row=sum_row, column=4)
    c.value = f'=SUM(F{DATA_START}:F{DATA_START+30})'; c.number_format = '¥#,##0.00'
    style_formula(c, bold=True, color=C_GREEN_DARK)

    # 店员A当月实发
    ws.cell(row=sum_row+1, column=1, value='店员A当月应发').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=sum_row+1, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=sum_row+1, column=1).border = thin_border()
    ws.cell(row=sum_row+1, column=1).fill = PatternFill('solid', start_color=C_PURPLE_BG)
    c = ws.cell(row=sum_row+1, column=2)
    c.value = (
        f'=SUM(J{DATA_START}:J{DATA_START+30})'
        f'+SUM(G{DATA_START}:G{DATA_START+30})'  # 人均提成之和（仅当班日）
        f'+SUM(H{DATA_START}:H{DATA_START+30})'  # 大单奖励
    )
    c.number_format = '¥#,##0.00'
    c.font = Font(name='Arial', bold=True, size=13, color='FF1A237E')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_BLUE_BG); c.border = med_border()
    ws.row_dimensions[sum_row+1].height = 28

    # 店员B当月实发
    ws.cell(row=sum_row+2, column=1, value='店员B当月应发').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=sum_row+2, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=sum_row+2, column=1).border = thin_border()
    ws.cell(row=sum_row+2, column=1).fill = PatternFill('solid', start_color=C_PURPLE_BG)
    c = ws.cell(row=sum_row+2, column=2)
    c.value = (
        f'=SUM(K{DATA_START}:K{DATA_START+30})'
        f'+SUM(G{DATA_START}:G{DATA_START+30})'
        f'+SUM(I{DATA_START}:I{DATA_START+30})'
    )
    c.number_format = '¥#,##0.00'
    c.font = Font(name='Arial', bold=True, size=13, color='FF1A237E')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_BLUE_BG); c.border = med_border()
    ws.row_dimensions[sum_row+2].height = 28

    # 注释
    ws.merge_cells(f'C{sum_row+1}:L{sum_row+2}')
    note = ws[f'C{sum_row+1}']
    note.value = (
        '⚠️ 说明：当月应发 = Σ日薪 + Σ人均提成（仅当班日计入）+ 大单奖励\n'
        '人均提成列(G列)已按当班人数平分。两人同班时两人各拿一份，单人班全拿。\n'
        '大单奖励按成交人归属（H列=店员A，I列=店员B），不参与平分。'
    )
    note.font = Font(name='Arial', size=9, color='FF5D4037', italic=True)
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note.fill = PatternFill('solid', start_color=C_GOLD_BG)
    note.border = thin_border()
    ws.row_dimensions[sum_row+1].height = 44
    ws.row_dimensions[sum_row+2].height = 28

    ws.freeze_panes = f'A{DATA_START}'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 4 - 库存管理
# ══════════════════════════════════════════════════════════════════════
def build_inventory_sheet(wb, df):
    ws = wb.create_sheet('📊库存管理')
    set_title(ws, '📊 库存管理  |  月度盘点', 'L', row=1)
    set_note(ws, '💡 期末 = 期初 + 入库 - 销售 | 库存≤2件标黄预警，0件标红', 'L', row=2)

    ws['A3'].value = '统计月份：'; ws['A3'].font = Font(name='Arial', bold=True, size=10)
    ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B3'].value = '2026-04'; style_input(ws['B3'])
    ws.row_dimensions[3].height = 22

    headers = ['序号','货号(SKU)','商品名称','颜色','尺码','售价(¥)','期初库存','本月入库','本月销售','期末库存','库存状态','备注']
    widths  = [5,14,32,18,7,12,10,10,10,10,12,16]
    draw_headers(ws, headers, widths, row=4)

    def split_cs(val):
        if pd.isna(val): return '', ''
        p = str(val).split(';')
        return p[0].strip(), (p[1].strip() if len(p) > 1 else '')
    df2 = df.copy()
    df2[['颜色','尺码']] = df2['线上颜色规格'].apply(lambda x: pd.Series(split_cs(x)))

    for ri, (_, row) in enumerate(df2.iterrows(), 5):
        ws.row_dimensions[ri].height = 17
        bg = C_ALT_ROW if ri % 2 == 0 else 'FFFFFFFF'
        fill = PatternFill('solid', start_color=bg)

        def bc(col, val, align='center', fmt=None, fill_override=None):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.border = thin_border()
            c.fill = fill_override or fill
            if fmt: c.number_format = fmt

        bc(1, ri-4); bc(2, str(row['原始商品编码']) if pd.notna(row['原始商品编码']) else '', 'center')
        bc(3, str(row['线上商品名称']) if pd.notna(row['线上商品名称']) else '', 'left')
        bc(4, row['颜色'], 'left'); bc(5, row['尺码'])
        bc(6, row['店铺售价'] if pd.notna(row['店铺售价']) else '', fmt='¥#,##0')

        # 期初（输入）
        c = ws.cell(row=ri, column=7, value=int(row['店铺库存']) if pd.notna(row['店铺库存']) else 0)
        c.font = Font(name='Arial', size=9); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = PatternFill('solid', start_color=C_INPUT_BG)

        # 入库（输入）
        c = ws.cell(row=ri, column=8, value=0)
        c.font = Font(name='Arial', size=9); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = PatternFill('solid', start_color=C_INPUT_BG)

        # 销售（输入）
        c = ws.cell(row=ri, column=9, value=0)
        c.font = Font(name='Arial', size=9); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = PatternFill('solid', start_color=C_INPUT_BG)

        # 期末（公式）
        c = ws.cell(row=ri, column=10)
        c.value = f'=G{ri}+H{ri}-I{ri}'
        c.font = Font(name='Arial', size=10, bold=True); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = PatternFill('solid', start_color=C_FORMULA_BG)

        # 状态
        c = ws.cell(row=ri, column=11)
        c.value = f'=IF(J{ri}=0,"🔴缺货",IF(J{ri}<=2,"🟡低库存","🟢正常"))'
        c.font = Font(name='Arial', size=9); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = PatternFill('solid', start_color=C_FORMULA_BG)

        bc(12, '', fill_override=fill)

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:L{len(df2)+4}'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 5 - 入库记录
# ══════════════════════════════════════════════════════════════════════
def build_inbound_sheet(wb):
    ws = wb.create_sheet('📥入库记录')
    set_title(ws, '📥 入库记录  |  每次到货登记', 'H', row=1)
    set_note(ws, '💡 每次到货后在此记录 | 填完后手动更新【库存管理】的本月入库列', 'H', row=2)

    headers = ['序号','入库日期','货号(SKU)','商品名称','颜色规格','入库数量','供应商/备注','经手人']
    widths  = [5,14,14,36,22,12,22,12]
    draw_headers(ws, headers, widths, row=3)

    for r in range(4, 204):
        ws.row_dimensions[r].height = 18
        bg = C_ALT_ROW if r % 2 == 0 else 'FFFFFFFF'

        c = ws.cell(row=r, column=1)
        c.value = f'=IF(B{r}="","",ROW()-3)'
        c.font = Font(name='Arial', size=9, color='FF9E9E9E')
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin_border()
        c.fill = PatternFill('solid', start_color=bg)

        for col in range(2, 9):
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill('solid', start_color=C_INPUT_BG)
            c.border = thin_border(); c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='center' if col != 4 else 'left', vertical='center')
        ws.cell(row=r, column=2).number_format = 'YYYY-MM-DD'

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = 'A3:H203'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 6 - 月度汇总
# ══════════════════════════════════════════════════════════════════════
def build_monthly_summary(wb):
    ws = wb.create_sheet('📈月度汇总')
    set_title(ws, '📈 月度汇总  |  全年经营一览', 'J', row=1)

    headers = ['月份','销售天数','月总销售额(¥)','销售件数','店员A工资(¥)','店员B工资(¥)','提成合计(¥)','经营达标情况','备注']
    widths  = [12,10,16,10,16,16,14,18,20]
    draw_headers(ws, headers, widths, row=2)

    months = [f'2026-{m:02d}' for m in range(1, 13)]
    for r, m in enumerate(months, 3):
        ws.row_dimensions[r].height = 22
        bg = C_ALT_ROW if r % 2 == 0 else 'FFFFFFFF'
        fill = PatternFill('solid', start_color=bg)

        c = ws.cell(row=r, column=1, value=m)
        c.font = Font(name='Arial', bold=True, size=10); c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border(); c.fill = fill

        for col in range(2, 10):
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill('solid', start_color=C_INPUT_BG); c.border = thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name='Arial', size=10)
            if col in [3, 5, 6, 7]: c.number_format = '¥#,##0'

    ws.freeze_panes = 'A3'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  Sheet 7 - 补货导出
# ══════════════════════════════════════════════════════════════════════
def build_replenish_sheet(wb, df):
    ws = wb.create_sheet('🚚补货导出')

    set_title(ws, '🚚 补货导出  |  基于库存自动生成补货清单，发给仓库补货', 'K', row=1)
    set_note(ws, '💡 补货数量 = 建议库存 - 期末库存（自动算）| 建议库存可修改 | 复制数据区发给仓库即可', 'K', row=2)

    headers = ['序号','货号(SKU)','款式编码','商品名称','颜色','尺码','售价(¥)',
               '当前库存','建议库存','补货数量','图片链接','供应商/备注']
    widths  = [5,14,12,36,18,7,11,10,10,10,42,22]
    draw_headers(ws, headers, widths, row=3)

    def split_cs(val):
        if pd.isna(val): return '', ''
        p = str(val).split(';')
        return p[0].strip(), (p[1].strip() if len(p) > 1 else '')

    df2 = df.copy()
    df2[['颜色','尺码']] = df2['线上颜色规格'].apply(lambda x: pd.Series(split_cs(x)))

    for ri, (_, row) in enumerate(df2.iterrows(), 4):
        ws.row_dimensions[ri].height = 17
        bg = C_ALT_ROW if ri % 2 == 0 else 'FFFFFFFF'
        fill = PatternFill('solid', start_color=bg)

        def bc(col, val, align='center', fmt=None, fill_override=None):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.border = thin_border()
            c.fill = fill_override or fill
            if fmt: c.number_format = fmt

        bc(1, ri-3)
        bc(2, str(row['原始商品编码']) if pd.notna(row['原始商品编码']) else '')
        bc(3, str(row['线上款式编码']) if pd.notna(row['线上款式编码']) else '')
        bc(4, str(row['线上商品名称']) if pd.notna(row['线上商品名称']) else '', 'left')
        bc(5, row['颜色'], 'left')
        bc(6, row['尺码'])
        bc(7, row['店铺售价'] if pd.notna(row['店铺售价']) else '', fmt='¥#,##0')

        # 当前库存（引用库存管理表的期末库存）
        bc(8, int(row['店铺库存']) if pd.notna(row['店铺库存']) else 0,
           fill_override=PatternFill('solid', start_color=C_FORMULA_BG))

        # 建议库存（输入，默认3）
        c = ws.cell(row=ri, column=9, value=3)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)

        # 补货数量（公式 = 建议库存 - 当前库存，最小0）
        c = ws.cell(row=ri, column=10)
        c.value = f'=MAX(I{ri}-H{ri},0)'
        c.font = Font(name='Arial', size=10, bold=True, color=C_GREEN_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        c.fill = PatternFill('solid', start_color=C_FORMULA_BG)

        # 图片链接
        img_url = str(row['图片']) if pd.notna(row['图片']) else ''
        c = ws.cell(row=ri, column=11, value=img_url)
        if img_url.startswith('http'):
            c.hyperlink = img_url
            c.font = Font(name='Arial', size=8, color='FF1565C0', underline='single')
        else:
            c.font = Font(name='Arial', size=8)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = thin_border(); c.fill = fill

        # 供应商/备注（输入）
        c = ws.cell(row=ri, column=12)
        c.fill = PatternFill('solid', start_color=C_INPUT_BG)
        c.border = thin_border(); c.font = Font(name='Arial', size=9)

    # 汇总区
    last_data = len(df2) + 3
    sum_row = last_data + 2
    ws.row_dimensions[sum_row].height = 24
    ws.merge_cells(f'A{sum_row}:G{sum_row}')
    c = ws[f'A{sum_row}']; c.value = '📦 需补货SKU总数'
    c.font = Font(name='Arial', bold=True, size=11, color=C_HEADER_FG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_HEADER_BG); c.border = thin_border()

    c = ws.cell(row=sum_row, column=8)
    c.value = f'=COUNTIF(J4:J{last_data},">0")'
    c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF00')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color='FF1A237E'); c.border = thin_border()

    ws.row_dimensions[sum_row+1].height = 24
    ws.merge_cells(f'A{sum_row+1}:G{sum_row+1}')
    c = ws[f'A{sum_row+1}']; c.value = '📦 补货总件数'
    c.font = Font(name='Arial', bold=True, size=11, color=C_HEADER_FG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color=C_HEADER_BG); c.border = thin_border()

    c = ws.cell(row=sum_row+1, column=8)
    c.value = f'=SUM(J4:J{last_data})'
    c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF00')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', start_color='FF1A237E'); c.border = thin_border()

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:L{last_data}'
    return ws


# ══════════════════════════════════════════════════════════════════════
#  导出飞书多维表格CSV
# ══════════════════════════════════════════════════════════════════════
def export_feishu_csv(df):
    """生成飞书多维表格可导入的CSV，含图片URL列"""
    def split_cs(val):
        if pd.isna(val): return '', ''
        p = str(val).split(';')
        return p[0].strip(), (p[1].strip() if len(p) > 1 else '')

    df2 = df.copy()
    df2[['颜色','尺码']] = df2['线上颜色规格'].apply(lambda x: pd.Series(split_cs(x)))

    # 只导出需要的列，图片链接单独一列（飞书导入后可设为附件字段）
    out = df2[['原始商品编码','线上款式编码','线上商品名称','颜色','尺码','店铺售价','店铺库存','图片']].copy()
    out.columns = ['货号(SKU)','款式编码','商品名称','颜色','尺码','售价(¥)','库存','图片链接']

    csv_path = OUTPUT_FILE.replace('.xlsx', '_飞书导入.csv')
    out.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f'  → 飞书导入CSV已保存至：{csv_path}')
    return csv_path


# ══════════════════════════════════════════════════════════════════════
#  主程序
# ══════════════════════════════════════════════════════════════════════
def main():
    print('读取产品数据...')
    df = pd.read_excel(SOURCE_FILE)
    print(f'  → {len(df)} 条SKU，{df["线上款式编码"].nunique()} 个款式')

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print('① 构建产品库...')
    build_product_sheet(wb, df)

    print('② 构建每日销售...')
    build_daily_sales_sheet(wb, df)

    print('③ 构建薪资计算...')
    build_salary_sheet(wb)

    print('④ 构建库存管理...')
    build_inventory_sheet(wb, df)

    print('⑤ 构建入库记录...')
    build_inbound_sheet(wb)

    print('⑥ 构建月度汇总...')
    build_monthly_summary(wb)

    print('⑦ 构建补货导出...')
    build_replenish_sheet(wb, df)

    wb.save(OUTPUT_FILE)
    print(f'\n✅ Excel完成！文件已保存至：{OUTPUT_FILE}')

    print('\n⑧ 导出飞书多维表格CSV...')
    csv_path = export_feishu_csv(df)
    print(f'\n✅ 全部完成！')

if __name__ == '__main__':
    main()
